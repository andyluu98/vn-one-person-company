from pathlib import Path
from core.obsidian.doc_writer import DocWriter


def test_write_docx_from_markdown_template(tmp_path):
    template = tmp_path / "tpl.md"
    template.write_text("# {{title}}\n\n## Section\n\nContent {{value}}", encoding="utf-8")

    writer = DocWriter(output_root=tmp_path / "out")
    out = writer.write_docx(template, "doc1.docx", {"title": "My Plan", "value": "X"})

    assert out.exists()
    assert out.suffix == ".docx"

    from docx import Document
    doc = Document(str(out))
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "Content X" in text


def test_write_xlsx_with_rows(tmp_path):
    template = tmp_path / "tpl.xlsx"

    writer = DocWriter(output_root=tmp_path / "out")
    out = writer.write_xlsx(template, "tracker.xlsx", [
        {"date": "2026-05-06", "metric": "CTR", "value": 2.3},
        {"date": "2026-05-07", "metric": "CTR", "value": 2.5},
    ])

    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.cell(1, 1).value == "date"
    assert ws.cell(2, 3).value == 2.3
